from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook


def build_mrv_template_xlsx() -> bytes:
    """
    MRV Excel Template (Streamlit Cloud uyumlu).

    Sheet isimleri (deterministik):
      - energy
      - production
      - materials
      - cbam_defaults  (Step-3: DEFAULT intensiteler + kanıt alanları)
    """
    wb = Workbook()

    ws = wb.active
    ws.title = "README"
    ws.append(["CME Demo — MRV Excel Template (Faz 3)"])
    ws.append([""])
    ws.append(["Sheet isimleri: energy, production, materials, cbam_defaults"])
    ws.append(["Kolon adlarını mümkünse değiştirmeyin. Tarih formatı: YYYY-MM"])
    ws.append([""])
    ws.append(["energy kolonları: month, facility_id, fuel_type, fuel_quantity, fuel_unit"])
    ws.append(["production kolonları: month, facility_id, sku, cn_code, quantity, unit, export_to_eu_quantity, actual_default_flag (opsiyonel)"])
    ws.append(["materials kolonları (2 mod):"])
    ws.append(["  (A) Basit EF: month, facility_id, material, quantity, unit, emission_factor"])
    ws.append(["  (B) Precursor zinciri: sku, precursor_sku, precursor_quantity, unit, (opsiyonel) precursor_embedded_tco2"])
    ws.append(["cbam_defaults kolonları: cbam_good_key, cn_code (opsiyonel), direct_intensity_tco2_per_unit, indirect_intensity_tco2_per_unit, unit, source, version, valid_from, valid_to, priority"])
    ws.append([""])

    ws_e = wb.create_sheet("energy")
    ws_e.append(["month", "facility_id", "fuel_type", "fuel_quantity", "fuel_unit"])
    ws_e.append(["2025-01", "1", "natural_gas", "1000", "Nm3"])
    ws_e.append(["2025-01", "1", "electricity", "50000", "kWh"])

    ws_p = wb.create_sheet("production")
    ws_p.append(["month", "facility_id", "sku", "cn_code", "quantity", "unit", "export_to_eu_quantity", "actual_default_flag"])
    ws_p.append(["2025-01", "1", "SKU-A", "7207", "1000", "t", "200", "ACTUAL"])
    ws_p.append(["2025-01", "1", "SKU-B", "2523", "500", "t", "100", "DEFAULT"])

    ws_m = wb.create_sheet("materials")
    ws_m.append(["sku", "precursor_sku", "precursor_quantity", "unit", "precursor_embedded_tco2"])
    ws_m.append(["SKU-A", "SKU-INPUT-1", "10", "t", "0.0"])

    ws_d = wb.create_sheet("cbam_defaults")
    ws_d.append(
        [
            "cbam_good_key",
            "cn_code",
            "direct_intensity_tco2_per_unit",
            "indirect_intensity_tco2_per_unit",
            "unit",
            "source",
            "version",
            "valid_from",
            "valid_to",
            "priority",
        ]
    )
    ws_d.append(["cement", "2523", "0.650", "0.050", "t", "Örnek kaynak", "v1", "2025-01-01", "2025-12-31", "10"])
    ws_d.append(["iron_steel", "", "2.100", "0.150", "t", "Örnek kaynak", "v1", "2025-01-01", "2025-12-31", "5"])

    for w in [ws_e, ws_p, ws_m, ws_d]:
        w.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
